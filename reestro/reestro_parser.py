# -*- coding: utf-8 -*-
"""
Парсер открытых сведений ЕГРН через API Контур.Реестро (realty.address-api v2).

Входной файл: TZ/Запрос.xlsx (или любой XLSX/CSV с подходящими столбцами).
  Поддерживаемые форматы:
    A) Запрос.xlsx — 16 колонок: Регион…ФИАС…ЕГРН…Адрес полностью
    B) Произвольный CSV/XLSX — столбцы «кадастровый номер» и/или «адрес»

Стратегия поиска кадастрового номера:
  Только по столбцу «ЕГРН» / «кадастровый номер».
  Объекты без КН пропускаются без запроса к API.

Выход:
  PDF — один файл на объект с валидным КН, формат по образцу ТЗ (Раздел 1 + Раздел 2)
  report.xlsx — сводная таблица (одна строка на каждое право), 35 колонок
  Статистика — лист «Сводка» в report.xlsx

Строки без кадастрового номера попадают только в Excel; PDF по ним не формируется.
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import uuid
from datetime import datetime, date
from pathlib import Path

import requests
from requests.exceptions import RequestException

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:
    print("Не установлен openpyxl. Выполните: pip install -r requirements.txt", file=sys.stderr)
    raise

try:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
except ImportError:
    print("Не установлен fpdf2. Выполните: pip install -r requirements.txt", file=sys.stderr)
    raise


# --------------------------------------------------------------------------- #
# Константы
# --------------------------------------------------------------------------- #
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_BASE_URL = "https://api.kontur.ru"

FONT_CANDIDATES = [
    (r"C:\Windows\Fonts\times.ttf", r"C:\Windows\Fonts\timesbd.ttf"),
    (r"C:\Windows\Fonts\arial.ttf",  r"C:\Windows\Fonts\arialbd.ttf"),
    (r"C:\Windows\Fonts\segoeui.ttf", r"C:\Windows\Fonts\segoeuib.ttf"),
    (r"C:\Windows\Fonts\calibri.ttf", r"C:\Windows\Fonts\calibrib.ttf"),
    ("/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
     "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf"),
    ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
     "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
]

ADDRESS_HEADERS = {"адрес", "address", "адрес объекта", "адрес объекта недвижимости"}
CADASTRAL_HEADERS = {
    "кадастровый номер", "кадастровый номер (при наличии)", "кн",
    "cadastral", "cadastral_number", "cadastralnumber", "кадастр", "егрн",
}
FIAS_HEADERS = {"фиас", "fias", "guid фиас", "фиас guid"}
EXT_HEADERS = {"доп. информация", "ext_number", "ext number"}

CADASTRAL_RE = re.compile(r"\d{2}:\d{2}:\d{1,7}:\d+")

# Образец: 43:31:070604:232-43/052/2023-24
RIGHT_NUM_RE  = re.compile(r"№\s*([\w:\/\-\.]+)")
RIGHT_DATE_RE = re.compile(r"от\s*(\d{2}\.\d{2}\.\d{4})")
RIGHT_SHARE_RE = re.compile(r"\b(\d+/\d+)\b")
RIGHT_TYPE_PAREN_RE = re.compile(r"\(([^)]+)\)")


# --------------------------------------------------------------------------- #
# Вспомогательные функции
# --------------------------------------------------------------------------- #
def _norm(s) -> str:
    return str(s or "").strip().lower()


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
    return str(v).strip()


def sanitize_filename(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|\r\n]+", "_", name).strip(" ._")
    return (name or "object")[:150]


def new_extract_id() -> str:
    """Номер выписки (UUID), как guidgenerator.ru."""
    return str(uuid.uuid4())


def pdf_name_for_extract(extract_id: str) -> str:
    return f"{extract_id}.pdf" if extract_id else ""


def _register_font(pdf: "FPDF") -> str:
    for regular, bold in FONT_CANDIDATES:
        if os.path.exists(regular):
            pdf.add_font("Body", "", regular)
            if bold and os.path.exists(bold):
                pdf.add_font("Body", "B", bold)
            else:
                pdf.add_font("Body", "B", regular)
            return "Body"
    raise RuntimeError(
        "Не найден TTF-шрифт с поддержкой кириллицы. "
        "Установите Arial / DejaVuSans или пропишите путь в FONT_CANDIDATES."
    )


# --------------------------------------------------------------------------- #
# Конфигурация / авторизация
# --------------------------------------------------------------------------- #
def load_config(config_path: Path) -> dict:
    cfg = {"baseUrl": DEFAULT_BASE_URL, "apiKey": "", "orgId": ""}
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            cfg.update({k: v for k, v in json.load(f).items() if v})
    cfg["apiKey"] = os.environ.get("REESTRO_API_KEY", cfg.get("apiKey", ""))
    cfg["orgId"]  = os.environ.get("REESTRO_ORG_ID",  cfg.get("orgId", ""))
    cfg["baseUrl"]= os.environ.get("REESTRO_BASE_URL", cfg.get("baseUrl", DEFAULT_BASE_URL))
    if not cfg["apiKey"] or not cfg["orgId"]:
        raise SystemExit(
            "Не заданы apiKey/orgId. Укажите в config.json или в переменных окружения "
            "REESTRO_API_KEY / REESTRO_ORG_ID."
        )
    return cfg


def auth_header(cfg: dict) -> dict:
    return {"Authorization": f"ReestroAuth apiKey={cfg['apiKey']}&portal.orgid={cfg['orgId']}"}


# --------------------------------------------------------------------------- #
# API-клиент
# --------------------------------------------------------------------------- #
class ReestroClient:
    def __init__(self, cfg: dict, pause: float = 0.4, timeout: int = 120,
                 retries: int = 5):
        self.base = cfg["baseUrl"].rstrip("/")
        self.session = requests.Session()
        self.session.headers.update(auth_header(cfg))
        proxy = cfg.get("proxy") or os.environ.get("HTTPS_PROXY") or os.environ.get("HTTP_PROXY")
        if proxy:
            self.session.proxies.update({"http": proxy, "https": proxy})
        self.pause = pause
        self.timeout = timeout
        self.retries = max(1, retries)

    def _get(self, url: str, params: dict | None = None):
        last_exc: RequestException | None = None
        resp = None
        for attempt in range(self.retries):
            try:
                resp = self.session.get(url, params=params, timeout=self.timeout)
            except RequestException as exc:
                last_exc = exc
                if attempt + 1 >= self.retries:
                    break
                wait = min(30, 3 * (attempt + 1))
                print(f"\n    сеть: {exc.__class__.__name__}, "
                      f"повтор {attempt + 2}/{self.retries} через {wait}с ...",
                      flush=True)
                time.sleep(wait)
                continue
            if resp.status_code == 429:
                if attempt + 1 >= self.retries:
                    break
                wait = min(30, 3 * (attempt + 1))
                print(f"\n    HTTP 429, повтор через {wait}с ...", flush=True)
                time.sleep(wait)
                continue
            time.sleep(self.pause)
            return resp
        if last_exc is not None:
            raise last_exc
        return resp

    def object_info(self, cadastral: str):
        """GET /realty/address/v2/objects/{cadastralNumber} → EstateObjectInfo."""
        return self._get(f"{self.base}/realty/address/v2/objects/{cadastral}")

    # Поиск по адресу отключен по требованиям


# --------------------------------------------------------------------------- #
# Чтение входного файла
# --------------------------------------------------------------------------- #
class InputRow:
    """Одна строка входного файла с нормализованными полями."""
    __slots__ = (
        "ext_number",           # «Доп. Информация» или порядковый ID
        "region_type",
        "region_name",
        "district_type",
        "district_name",
        "city_type",
        "city_name",
        "locality_type",
        "locality_name",
        "street_type",
        "street_name",
        "house_type",
        "house_num",
        "korpus_type",
        "korpus_num",
        "stroenie_type",
        "stroenie_num",
        "pomesh_type",
        "pomesh_num",
        "fias_guid",
        "cadastral",
        "full_address",
    )

    def __init__(self):
        for f in self.__slots__:
            setattr(self, f, "")


def _read_zapros_xlsx(path: Path) -> list[InputRow]:
    """Читает файл в формате TZ/Запрос.xlsx (16 + специальных колонок)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hdrs = [_norm(h) for h in rows[0]]
    MAPPING = {
        "регион":                    "region_name",
        "тип района":                "district_type",
        "название района":           "district_name",
        "тип населенного пункта":    "locality_type",
        "название населенного пункта":"locality_name",
        "тип улицы":                 "street_type",
        "название улицы":            "street_name",
        "дом":                       "house_num",
        "корпус":                    "korpus_num",
        "строение":                  "stroenie_num",
        "помещение":                 "pomesh_num",
        "комната":                   "pomesh_num",   # fallback
        "доп. информация":           "ext_number",
        "доп. информация":           "ext_number",
        "фиас":                      "fias_guid",
        "егрн":                      "cadastral",
        "адрес полностью":           "full_address",
    }

    col_map: dict[int, str] = {}
    for i, h in enumerate(hdrs):
        field = MAPPING.get(h)
        if field:
            col_map[i] = field

    result = []
    for raw in rows[1:]:
        if not any(_norm(c) for c in raw):
            continue
        row = InputRow()
        for i, field in col_map.items():
            if i < len(raw):
                setattr(row, field, _str(raw[i]))
        result.append(row)
    return result


def _read_generic(path: Path) -> list[InputRow]:
    """Читает CSV или XLSX с произвольными столбцами address/cadastral."""
    raw_rows: list[list] = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        for r in wb.active.iter_rows(values_only=True):
            raw_rows.append(list(r))
    else:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            except csv.Error:
                dialect = csv.excel
                dialect.delimiter = ";"
            for r in csv.reader(f, dialect):
                raw_rows.append(r)

    raw_rows = [r for r in raw_rows if any(_norm(c) for c in r)]
    if not raw_rows:
        raise SystemExit("Входной файл пуст.")

    hdrs = [_norm(h) for h in raw_rows[0]]
    addr_idx = cad_idx = fias_idx = ext_idx = None
    for i, h in enumerate(hdrs):
        if addr_idx is None and h in ADDRESS_HEADERS:
            addr_idx = i
        if cad_idx is None and h in CADASTRAL_HEADERS:
            cad_idx = i
        if fias_idx is None and h in FIAS_HEADERS:
            fias_idx = i
        if ext_idx is None and h in EXT_HEADERS:
            ext_idx = i

    data = raw_rows[1:] if (addr_idx is not None or cad_idx is not None) else raw_rows
    if addr_idx is None and cad_idx is None:
        addr_idx, cad_idx = 0, 1

    result = []
    for r in data:
        def c(i):
            return _str(r[i]) if (i is not None and i < len(r)) else ""
        row = InputRow()
        row.full_address = c(addr_idx)
        row.cadastral    = c(cad_idx)
        row.fias_guid    = c(fias_idx) if fias_idx is not None else ""
        row.ext_number   = c(ext_idx) if ext_idx is not None else ""
        if row.full_address or row.cadastral:
            result.append(row)
    return result


def read_input(path: Path) -> list[InputRow]:
    """Определяет формат файла и читает его."""
    if path.suffix.lower() not in (".xlsx", ".xlsm", ".csv"):
        raise SystemExit(f"Неподдерживаемый формат: {path.suffix}")

    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        first_row = next(wb.active.iter_rows(values_only=True), ())
        hdrs = {_norm(h) for h in first_row if h}
        if "фиас" in hdrs or "адрес полностью" in hdrs or "егрн" in hdrs:
            return _read_zapros_xlsx(path)

    return _read_generic(path)


# --------------------------------------------------------------------------- #
# Разбор прав из ответа API
# --------------------------------------------------------------------------- #
def parse_right_string(right_str: str) -> dict:
    """
    Разбирает строку вида:
      «№ 43:31:070604:232-43/052/2023-24 от 26.10.2023 (Долевая собственность, 1/5)»
    Возвращает словарь {right_type, share, number, reg_date, full}.

    Доля извлекается только из скобочного блока, чтобы не захватить части
    кадастрового номера вроде «43/052».
    """
    r = right_str or ""
    result = {"right_type": "", "share": "", "number": "", "reg_date": "", "full": r}

    m = RIGHT_NUM_RE.search(r)
    if m:
        result["number"] = m.group(1)

    m = RIGHT_DATE_RE.search(r)
    if m:
        result["reg_date"] = m.group(1)

    # Вид права и доля — из содержимого скобок: «(Долевая собственность, 1/5)»
    m = RIGHT_TYPE_PAREN_RE.search(r)
    if m:
        inner = m.group(1)
        parts = [p.strip() for p in inner.split(",")]
        result["right_type"] = parts[0]
        if len(parts) > 1:
            # Доля вида «1/5», «2/3» — только простые дроби, не часть КН
            candidate = parts[1].strip()
            if re.fullmatch(r"\d{1,3}/\d{1,4}", candidate):
                result["share"] = candidate
    elif r and not result["number"]:
        # Строка — просто название права без регномера
        result["right_type"] = r.strip("() ")

    return result


def extract_rights(info: dict) -> list[dict]:
    """
    Из EstateObjectInfo.rightAndRestrictions строит список прав:
    каждый элемент = {owner_type, right_type, share, number, reg_date,
                      encumbrances, full_right}.
    """
    rights = []
    owner_type_global = info.get("ownershipType") or ""

    for ri in (info.get("rightAndRestrictions") or []):
        right_raw = (ri.get("right") or "").strip()
        parsed    = parse_right_string(right_raw)

        encs = []
        for enc in (ri.get("restrictions") or []):
            if enc:
                encs.append(str(enc).strip())

        owner_type = (ri.get("rightOwnerType") or
                      ri.get("ownerType") or
                      owner_type_global)

        rights.append({
            "owner_type":   owner_type,
            "right_type":   parsed["right_type"],
            "share":        parsed["share"],
            "number":       parsed["number"],
            "reg_date":     parsed["reg_date"],
            "encumbrances": encs,
            "full_right":   right_raw,
        })

    return rights


# --------------------------------------------------------------------------- #
# Генерация PDF (Раздел 1 + Раздел 2)
# --------------------------------------------------------------------------- #
NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}


def _dn(value) -> str:
    """Возвращает значение либо 'данные отсутствуют', как в образце."""
    s = _str(value)
    return s if s else "данные отсутствуют"


def _format_floor(info: dict) -> str:
    """Форматирует этаж как в образце: «Этаж №03»."""
    floor = _str(info.get("floor") or "")
    floors = _str(info.get("floors") or "")
    if floor:
        num = floor if floor.lower().startswith("этаж") else f"Этаж №{floor}"
        if floors:
            return f"{num} / {floors}"
        return num
    if floors:
        return floors
    return ""


class EgrnPdf(FPDF):
    """
    PDF-отчёт строго по образцу заказчика (LibreOffice / Times New Roman):
      - альбомная ориентация Letter (792×612 пт);
      - таблицы с рамками, без выделения жирным;
      - колонка «метка» уже колонки «значение»;
      - раздел 2 — таблица из 4 колонок (№ | описание | под-№ | значение);
      - повтор шапки (вид объекта + лист № X + дата + КН) на каждой странице.
    Счётчики «Всего листов …» проставляются вторым проходом (см. generate_pdf).
    """

    def __init__(self, obj_type: str, cadastral: str, report_date: str,
                 info: dict, rights: list[dict], input_row: "InputRow",
                 totals: tuple | None = None, ownership_form: str = ""):
        super().__init__(orientation="L", unit="mm", format="letter")
        self.set_margins(9, 20, 9)              # боковые ~26 pt; верх 20 мм
        self.b_margin = 30                        # низ ~86 pt как в образце
        self.set_auto_page_break(auto=True, margin=30)
        self.c_margin = 0                         # без внутреннего отступа ячеек (как в образце)
        self._font = _register_font(self)

        self.obj_type  = obj_type or "Объект недвижимости"
        self.cadastral = cadastral
        self.rep_date  = report_date
        self._info     = info
        self._rights   = rights
        self._row      = input_row
        self._owner_form = _str(ownership_form)

        self._uw = self.w - self.l_margin - self.r_margin   # рабочая ширина
        self._indent = 20                                 # отступ 2 см для абзацев

        # Раздел 2: № | описание | под-№ | значение (30|250|30 pt в образце)
        self._c1 = 10.6
        self._c2 = 88.2
        self._c3 = 10.6
        self._c4 = self._uw - self._c1 - self._c2 - self._c3
        # Раздел 1 (и строка КН): метка = col1+col2, значение = col3+col4 — как в образце
        self._l1 = self._c1 + self._c2
        self._v1 = self._c3 + self._c4

        # высоты/зазоры по образцу (отдельные таблицы разделены белым полем)
        self._rh_row = 8.0          # строка «Помещение» / «вид…» / дата (~22 pt)
        self._rh_short = 5.1        # «Лист №» на листе 1 разд. 1 (~14 pt)
        self._rh_date = 8.0
        self._gap_blank = 5.0       # белое поле перед «Помещение» (лист 2+)
        self._gap_obj_list = 7.5      # белое поле «Помещение» → «Лист №» (~21 pt)
        self._gap_list_date = 7.5     # белое поле «Лист №» → дата (~21 pt)

        if totals:
            self.totals_s1, self.totals_s2, self.total_all = totals
        else:
            self.totals_s1 = self.totals_s2 = self.total_all = ""

        self.cur_section = 1
        self.section_first_page = 1
        self._s2_first = 2

    # ---- шрифт (без жирного — строго как в образце) ----
    def _f(self, size: int = 10):
        self.set_font(self._font, "", size)

    def _max_table_row_lines(self, line_height: float = 5.0) -> int:
        """Сколько строк текста помещается в одну строку таблицы fpdf2 на странице."""
        available = self.page_break_trigger - self.t_margin
        return max(1, int(available // line_height) - 2)

    def _wrap_text_lines(self, text: str, width: float) -> list[str]:
        """Перенос текста по ширине колонки (слова и сверхдлинные фрагменты без пробелов)."""
        self._f(10)
        if not text:
            return [""]

        lines: list[str] = []
        for para in text.replace("\r\n", "\n").split("\n"):
            if not para:
                lines.append("")
                continue

            line = ""
            for word in para.split():
                if not line:
                    candidate = word
                else:
                    candidate = f"{line} {word}"

                if self.get_string_width(candidate) <= width:
                    line = candidate
                    continue

                if line:
                    lines.append(line)
                    line = word
                else:
                    line = word

                while line and self.get_string_width(line) > width:
                    cut = 1
                    while cut < len(line) and self.get_string_width(line[:cut]) <= width:
                        cut += 1
                    cut = max(1, cut - 1)
                    lines.append(line[:cut])
                    line = line[cut:]

            if line:
                while self.get_string_width(line) > width:
                    cut = 1
                    while cut < len(line) and self.get_string_width(line[:cut]) <= width:
                        cut += 1
                    cut = max(1, cut - 1)
                    lines.append(line[:cut])
                    line = line[cut:]
                if line:
                    lines.append(line)

        return lines or [""]

    def _table_cell_chunks(self, text: str, width: float,
                           line_height: float = 5.0) -> list[str]:
        """Дробит текст ячейки на части, каждая из которых помещается на одну страницу."""
        lines = self._wrap_text_lines(text, width)
        max_lines = self._max_table_row_lines(line_height)
        chunks: list[str] = []
        for i in range(0, len(lines), max_lines):
            chunks.append("\n".join(lines[i:i + max_lines]))
        return chunks or [""]

    def _emit_section2_rows(self, table, num: str, label: str, sub_num: str,
                            value: str, *, line_height: float = 5.0,
                            value_colspan: int = 1):
        """
        Строка раздела 2; длинные label/value разбиваются на несколько строк таблицы,
        чтобы fpdf2 не падал с «row is too high».
        """
        label_chunks = self._table_cell_chunks(label, self._c2, line_height)
        if value_colspan == 2:
            value_chunks = self._table_cell_chunks(
                value, self._c3 + self._c4, line_height)
        else:
            value_chunks = self._table_cell_chunks(value, self._c4, line_height)

        row_count = max(len(label_chunks), len(value_chunks))
        for i in range(row_count):
            tr = table.row()
            tr.cell(num if i == 0 else "")
            tr.cell(label_chunks[i] if i < len(label_chunks) else "")
            if value_colspan == 2:
                tr.cell(value_chunks[i] if i < len(value_chunks) else "", colspan=2)
            else:
                tr.cell(sub_num if i == 0 else "")
                tr.cell(value_chunks[i] if i < len(value_chunks) else "")

    # ---- повторяющаяся шапка страницы ----
    def header(self):
        """Верх страницы: только вводный текст + шапка листа 1 (лист 2+ вставляется в теле)."""
        if self.page_no() != self.section_first_page:
            return

        sec = self.cur_section
        if sec == 1:
            self._f(9)
            self.multi_cell(self._uw, 4.5,
                "Отчёт сформирован согласно записям из ЕГРН, носит справочный "
                "характер и не\nявляется официальной выпиской",
                align="C", **NL)
            self.ln(5)
            self._f(11)
            self.multi_cell(self._uw, 6,
                "Сведения о характеристиках объекта недвижимости",
                align="C", **NL)
            self.ln(5)
            self._f(10)
            self.set_x(self.l_margin + self._indent)
            self.multi_cell(self._uw - self._indent, 5,
                f"На основании запроса от {self.rep_date}, поступившего на "
                f"рассмотрение {self.rep_date}, сообщаем, что согласно записям "
                "Единого государственного реестра\nнедвижимости:", **NL)
            self.set_x(self.l_margin)
            self._f(10)
            self.cell(self._uw, 4, "Раздел 1", align="R", **NL)
            self._repeat_header(1, 1, connected=True)
        else:
            self._f(10)
            self.cell(self._uw - self._indent, 5,
                "Раздел 2 Отчёт об объекте недвижимости", align="R", **NL)
            self._f(10)
            self.multi_cell(self._uw, 5,
                "Сведения о зарегистрированных правах", align="C", **NL)
            self._repeat_header(1, 2, connected=False)

    def footer(self):
        pass

    def _gap_white(self, h: float):
        """Белое поле между отдельными таблицами (без рамки)."""
        self.ln(h)

    def _cell_date_row(self, *, isolated: bool = True):
        """Строка даты. isolated=True — отдельная таблица после белого поля."""
        x = self.l_margin
        y = self.get_y()
        if isolated:
            self.rect(x, y, self._uw, self._rh_date, style="D")
        else:
            self.cell(self._uw, self._rh_date, "", border=1, align="L")
        self._f(10)
        self.text(x + 0.4, y + 2.8, f"{self.rep_date}г.")
        self.set_y(y + self._rh_date)

    def _draw_object_table(self):
        """Таблица 1: «Помещение» + «вид объекта недвижимости» (отдельный блок)."""
        self._f(11)
        self.cell(self._uw, self._rh_row, self.obj_type, border=1, align="C", **NL)
        self._f(10)
        self.cell(self._uw, self._rh_row, "вид объекта недвижимости",
                  border=1, align="C", **NL)

    def _draw_list_table(self, leaf: int, sec: int, *, short: bool = False):
        """Таблица 2: строка «Лист № …» (отдельный блок)."""
        total_s = self.totals_s1 if sec == 1 else self.totals_s2
        rh = self._rh_short if short else self._rh_row
        self._f(9)
        w = [60, 75, 56, self._uw - 191]
        self.cell(w[0], rh, f"Лист № {leaf} раздела {sec}",
                  border=1, align="C")
        self.cell(w[1], rh,
                  f"Всего листов раздела {sec}: {total_s}", border=1, align="C")
        self.cell(w[2], rh, "Всего разделов: 2", border=1, align="C")
        self.cell(w[3], rh,
                  f"Всего листов отчёта: {self.total_all}", border=1, align="C", **NL)

    def _draw_kn_row(self, *, short: bool = False):
        """Строка «Кадастровый номер» — первая строка основной таблицы данных."""
        rh = self._rh_short if short else self._rh_row
        self._f(10)
        self.cell(self._l1, rh, "Кадастровый номер:", border=1)
        self.cell(self._v1, rh, self.cadastral, border=1, **NL)

    def _draw_data_prefix(self, *, connected: bool = False):
        """Начало основной таблицы: дата + КН."""
        self._cell_date_row(isolated=not connected)
        self._draw_kn_row(short=connected)

    def _header_block_height(self, *, connected: bool, leaf: int) -> float:
        rh_list = self._rh_short if connected else self._rh_row
        gap_ld = 0 if connected else self._gap_list_date
        lead = self._gap_blank if leaf > 1 else 0
        return (lead + self._rh_row * 2 + self._gap_obj_list + rh_list + gap_ld)

    def _ensure_header_space(self, *, connected: bool, leaf: int,
                             extra: float = 0):
        """Не допускать разрыва блока «Помещение…Лист №» между страницами."""
        need = self._header_block_height(connected=connected, leaf=leaf) + extra
        if self.get_y() + need > self.page_break_trigger:
            self.add_page()

    def _repeat_header(self, leaf: int, sec: int, *, connected: bool = False):
        """Отдельные таблицы «Помещение» и «Лист №» с белыми полями (как в образце)."""
        if leaf > 1:
            self._gap_white(self._gap_blank)

        self._draw_object_table()
        self._gap_white(self._gap_obj_list)
        self._draw_list_table(leaf, sec, short=connected)

        if not connected:
            self._gap_white(self._gap_list_date)

    # ---- наполнение раздела 1 ----
    def _section1_fields(self) -> list[tuple]:
        info = self._info
        cad  = self.cadastral
        kvartal = ":".join(cad.split(":")[:3]) if cad else ""
        area   = _str(info.get("area") or "")
        a_unit = _str(info.get("areaUnit") or "")
        area_v = f"{area} {a_unit}".strip() if area else ""
        floor  = _format_floor(info)
        parent = info.get("parentCadastralNumbers") or []
        status = info.get("estateObjectStatus") or ""
        if "актуал" in status.lower():
            status = 'Сведения об объекте недвижимости имеют статус "Актуально"'
        notes = "; ".join(n for n in (info.get("specialNotes") or []) if n)

        return [
            ("Номер кадастрового квартала", kvartal),
            ("Дата присвоения кадастрового номера", info.get("createDate")),
            ("Ранее присвоенный государственный учетный номер",
             info.get("previousCadastralNumbers")),
            ("Адрес", info.get("address") or self._row.full_address),
            ("Площадь", area_v),
            ("Назначение", info.get("constructionType")),
            ("Наименование", info.get("realEstateGroup")),
            ("Номер, тип этажа, на котором расположено помещение, машино-место", floor),
            ("Вид жилого помещения", info.get("dwellingType")),
            ("Кадастровая стоимость, руб.", info.get("cadastralValue")),
            ("Кадастровые номера иных объектов недвижимости, в пределах которых "
             "расположен объект недвижимости", "; ".join(parent)),
            ("Кадастровые номера объектов недвижимости, из которых образован "
             "объект недвижимости", ""),
            ("Кадастровые номера образованных объектов недвижимости", ""),
            ("Сведения о включении объекта недвижимости в состав предприятия "
             "как имущественного комплекса", ""),
            ("Сведения о включении объекта недвижимости в состав единого "
             "недвижимого комплекса", ""),
            ("Виды разрешенного использования", info.get("permittedUse")),
            ("Сведения о включении объекта недвижимости в реестр объектов "
             "культурного наследия", ""),
            ("Сведения о кадастровом инженере", ""),
            ("Сведения о признании многоквартирного дома аварийным и подлежащим "
             "сносу или реконструкции и (или) о признании жилого помещения, в том "
             "числе жилого дома, непригодным для проживания", ""),
            ("Сведения об отнесении жилого помещения к определенному виду жилых "
             "помещений специализированного жилищного фонда, к жилым помещениям "
             "наемного дома социального использования или наемного дома "
             "коммерческого использования", ""),
            ("Статус записи об объекте недвижимости", status),
            ("Особые отметки", notes),
        ]

    def _section1_table(self, fields: list[tuple], *,
                        data_prefix: bool = False, connected: bool = False):
        if data_prefix:
            self._draw_data_prefix(connected=connected)
        if not fields:
            return
        self._f(10)
        with self.table(col_widths=(self._l1, self._v1), width=self._uw,
                        borders_layout="ALL", first_row_as_headings=False,
                        line_height=5, text_align=("LEFT", "LEFT"),
                        v_align="TOP", padding=0) as table:
            for label, value in fields:
                tr = table.row()
                lbl = label if label.endswith(":") else label + ":"
                tr.cell(lbl)
                tr.cell(_dn(value))

    def _section1_body(self):
        fields = self._section1_fields()
        split = next(
            i for i, (lbl, _) in enumerate(fields)
            if lbl.startswith("Сведения о кадастровом инженере")
        )
        self._section1_table(fields[:split],
                             data_prefix=True, connected=True)
        self._ensure_header_space(connected=False, leaf=2)
        self._repeat_header(2, 1, connected=False)
        self._section1_table(fields[split:],
                             data_prefix=True, connected=False)

    # ---- наполнение раздела 2 (зарегистрированные права) ----
    def _section2_table(self, rights: list[dict], global_items: list[tuple], *,
                        data_prefix: bool = False):
        if data_prefix:
            self._draw_data_prefix(connected=False)
        if not rights and not global_items:
            if not data_prefix:
                self._f(10)
                with self.table(col_widths=(self._c1, self._c2, self._c3, self._c4),
                                width=self._uw,
                                borders_layout="ALL", first_row_as_headings=False,
                                line_height=5,
                                text_align=("LEFT", "LEFT", "LEFT", "LEFT"),
                                v_align="TOP", padding=0) as table:
                    tr = table.row()
                    tr.cell("Сведения о зарегистрированных правах отсутствуют",
                            colspan=4)
            return
        self._f(10)
        with self.table(col_widths=(self._c1, self._c2, self._c3, self._c4),
                        width=self._uw,
                        borders_layout="ALL", first_row_as_headings=False,
                        line_height=5,
                        text_align=("LEFT", "LEFT", "LEFT", "LEFT"),
                        v_align="TOP", padding=0) as table:
            for r in rights:
                # «Правообладатель» в открытых сведениях = форма собственности
                # (с lk.rosreestr.ru), напр. «Государственная федеральная».
                owner = self._owner_form or "данные отсутствуют"
                self._emit_section2_rows(
                    table, "1", "Правообладатель (правообладатели):", "1.1", owner)

                self._emit_section2_rows(
                    table, "",
                    "Сведения о возможности предоставления третьим лицам "
                    "персональных данных физического лица:",
                    "1.1.1", "данные отсутствуют")

                right_type = r.get("right_type") or "данные отсутствуют"
                share = r.get("share") or ""
                right_line = right_type + (f", {share}" if share else "")
                val = right_line
                if r.get("number"):
                    val += f"\n{_str(r.get('number'))}"
                if r.get("reg_date"):
                    val += f"\n{_str(r.get('reg_date'))}"
                self._emit_section2_rows(
                    table, "2",
                    "Вид, номер, дата и время государственной регистрации права:",
                    "2.1", val)

                self._emit_section2_rows(
                    table, "3",
                    "Сведения об осуществлении государственной регистрации "
                    "сделки, права без необходимого в силу закона согласия "
                    "третьего лица, органа:",
                    "3.1", "данные отсутствуют")

                encs = r.get("encumbrances") or []
                enc_txt = "; ".join(encs) if encs else "не зарегистрировано"
                self._emit_section2_rows(
                    table, "4",
                    "Ограничение прав и обременение объекта недвижимости:",
                    "", enc_txt, value_colspan=2)

            for num, label, value in global_items:
                lbl = label if label.endswith(":") else label + ":"
                self._emit_section2_rows(
                    table, str(num), lbl, "", value, value_colspan=2)

    def _section2_body(self):
        rights = self._rights or []
        global_items = [
            ("5", "Договоры участия в долевом строительстве:",
             "не зарегистрировано"),
            ("6", "Заявленные в судебном порядке права требования:",
             "данные отсутствуют"),
            ("7", "Сведения о возможности предоставления третьим лицам "
                  "персональных данных физического лица",
             "данные отсутствуют"),
            ("8", "Сведения о возражении в отношении зарегистрированного "
                  "права:",
             "данные отсутствуют"),
            ("9", "Сведения о наличии решения об изъятии объекта "
                  "недвижимости для государственных и муниципальных нужд:",
             "данные отсутствуют"),
            ("10", "Сведения о невозможности государственной регистрации "
                   "без личного участия правообладателя или его законного "
                   "представителя:",
             "данные отсутствуют"),
            ("11", "Правопритязания и сведения о наличии поступивших, но не "
                   "рассмотренных заявлений о проведении государственной "
                   "регистрации права (перехода, прекращения права), "
                   "ограничения права или обременения объекта недвижимости, "
                   "сделки в отношении объекта недвижимости:",
             "данные отсутствуют"),
        ]
        split = next(i for i, (n, _, _) in enumerate(global_items) if n == "10")
        self._section2_table(rights, global_items[:split], data_prefix=True)
        extra = self._rh_date + self._rh_row + self._rh_row * 5
        self._ensure_header_space(connected=False, leaf=2, extra=extra)
        self._repeat_header(2, 2, connected=False)
        self._section2_table([], global_items[split:], data_prefix=True)

    # ---- сборка документа ----
    def build(self):
        self.cur_section = 1
        self.section_first_page = 1
        self.add_page()
        self._section1_body()

        self._s2_first = self.page_no() + 1
        self.cur_section = 2
        self.section_first_page = self._s2_first
        self.add_page()
        self._section2_body()


def generate_pdf(info: dict, rights: list[dict],
                 input_row: InputRow, out_path: Path,
                 ownership_form: str = ""):
    """Генерирует PDF-отчёт в формате образца (альбомный, Times New Roman)."""
    obj_type  = info.get("realEstateGroup") or "Объект недвижимости"
    cadastral = info.get("cadastralNumber") or input_row.cadastral or ""
    rep_date  = datetime.now().strftime("%d.%m.%Y")

    # 1-й проход — узнаём число листов в каждом разделе
    probe = EgrnPdf(obj_type, cadastral, rep_date, info, rights, input_row,
                    ownership_form=ownership_form)
    probe.build()
    s1_pages  = probe._s2_first - 1
    total_all = probe.page_no()
    s2_pages  = total_all - s1_pages

    # 2-й проход — с корректными счётчиками «Всего листов …»
    doc = EgrnPdf(obj_type, cadastral, rep_date, info, rights, input_row,
                  totals=(s1_pages, s2_pages, total_all),
                  ownership_form=ownership_form)
    doc.build()
    doc.output(str(out_path))


def generate_pdf_not_found(input_row: InputRow, cadastral: str,
                           reason: str, out_path: Path):
    """PDF-заглушка для объекта, по которому нет данных."""
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_margins(20, 20, 20)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    font = _register_font(pdf)
    pdf.set_font(font, "B", 13)
    pdf.multi_cell(0, 7,
        "Отчёт сформирован согласно записям из ЕГРН, "
        "носит справочный характер и не является официальной выпиской",
        **NL)
    pdf.ln(3)
    pdf.set_font(font, "", 10)
    addr = input_row.full_address or ""
    if addr:
        pdf.multi_cell(0, 6, f"Адрес из задания: {addr}", **NL)
    if input_row.fias_guid:
        pdf.multi_cell(0, 6, f"ФИАС GUID: {input_row.fias_guid}", **NL)
    if cadastral:
        pdf.multi_cell(0, 6, f"Кадастровый номер: {cadastral}", **NL)
    pdf.ln(3)
    pdf.set_text_color(180, 0, 0)
    pdf.set_font(font, "B", 10)
    pdf.multi_cell(0, 6,
        "Сведения по объекту недвижимости в ЕГРН получить не удалось.", **NL)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(font, "", 10)
    pdf.multi_cell(0, 6, f"Причина: {reason}", **NL)
    pdf.multi_cell(0, 6,
        f"Дата формирования: {datetime.now():%d.%m.%Y %H:%M}", **NL)
    pdf.output(str(out_path))


# --------------------------------------------------------------------------- #
# Генерация сводного Excel
# --------------------------------------------------------------------------- #
def _xlrow_base(input_row: InputRow) -> list:
    """A–C из Запрос.xlsx; D–AB (инд. 3–27) не заполняем."""
    row = [""] * 28
    row[0] = input_row.ext_number
    row[2] = input_row.full_address
    return row


def build_xlsx_rows(input_row: InputRow, info: dict | None,
                    rights: list[dict], pdf_filename: str,
                    extract_id: str, extract_date: str,
                    ownership_form: str = "") -> list[list]:
    """
    Строит строки для report.xlsx (одна строка = одно право объекта).
    A–C из Запрос.xlsx; D–AB пустые; права — из ответа API (открытые сведения ЕГРН).

    Колонки группы «Собственник» (0-индексы 28–31):
      28 «ФИО»               — открытые сведения ЕГРН не содержат ФИО → «данные отсутствуют»;
      29 «Вид собственности» — форма собственности с lk.rosreestr.ru (стр. «Форма собственности»);
      30 «Доля»              — доля в праве; пустая, если данных нет (правка заказчика);
      31 «Право»             — вид, номер и дата государственной регистрации права.
    """
    base = _xlrow_base(input_row)
    base[1] = input_row.cadastral

    # Форма собственности едина для объекта (берётся из Росреестра, не из API Контура).
    form = _str(ownership_form)
    pdf_filename = pdf_name_for_extract(extract_id) or pdf_filename

    def rights_block(fio, own_form, share, right) -> list:
        # «Доля» НЕ оборачиваем в _dn: пустая ячейка остаётся пустой.
        return [_dn(fio), _dn(own_form), _str(share), _dn(right)]

    result = []
    if not rights:
        row = base + rights_block("", form, "", "") + [
            extract_id, extract_date, pdf_filename,
        ]
        result.append(row)
    else:
        for r in rights:
            fio         = r.get("owner_type") or ""
            right_type  = r.get("right_type") or ""
            share       = r.get("share") or ""
            number      = r.get("number") or ""
            reg_date    = r.get("reg_date") or ""
            full        = r.get("full_right") or ""
            right_col   = right_type
            if number and reg_date:
                right_col = f"{right_type}, \u2116 {number} \u043e\u0442 {reg_date}"
            elif full:
                right_col = full

            row = base[:] + rights_block(
                fio, form, share, right_col,
            ) + [
                extract_id,
                extract_date,
                pdf_filename,
            ]
            result.append(row)
    return result


def normalize_kn(value) -> str:
    """Нормализованный кадастровый номер для сравнения."""
    s = _str(value)
    m = CADASTRAL_RE.search(s)
    return m.group(0) if m else s


def cache_json_path(cache_dir: Path, kn: str) -> Path:
    """Путь к файлу кэша ответа API для кадастрового номера."""
    return cache_dir / f"{sanitize_filename(normalize_kn(kn))}.json"


def save_api_cache(cache_dir: Path, kn: str, info: dict, resp,
                   input_row: "InputRow | None" = None) -> Path:
    """Сохраняет ответ API в JSON (промежуточные данные без повторных запросов)."""
    cache_dir.mkdir(parents=True, exist_ok=True)
    path = cache_json_path(cache_dir, kn)
    prev_input: dict = {}
    prev_payload: dict = {}
    if path.exists():
        try:
            with open(path, encoding="utf-8") as f:
                prev_payload = json.load(f)
                prev_input = prev_payload.get("input") or {}
        except (OSError, ValueError, json.JSONDecodeError):
            prev_input = {}
            prev_payload = {}

    inp = dict(prev_input)
    if input_row:
        if input_row.ext_number:
            inp["ext_number"] = input_row.ext_number
        if input_row.full_address:
            inp["full_address"] = input_row.full_address

    payload = {
        "cadastralNumber": normalize_kn(kn),
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "http_status": getattr(resp, "status_code", None) or prev_payload.get("http_status", 200),
        "request_id": _str(getattr(resp, "headers", {}).get("x-request-id", "")) or prev_payload.get("request_id", ""),
        "info": info,
    }
    if inp:
        payload["input"] = inp
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


def load_cache_payload(cache_dir: Path, kn: str) -> dict | None:
    """Полный JSON-кэш по КН (info + input + метаданные)."""
    path = cache_json_path(cache_dir, kn)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else None


# Входные файлы-справочники, не содержащие EXT_NUMBER — не подмешивать при авто-поиске.
INPUT_SKIP_NAMES = {"ownership_forms.csv"}


def input_rows_by_kn(path: Path) -> dict[str, InputRow]:
    """Индекс входных строк по кадастровому номеру."""
    if not path.exists():
        return {}
    if path.name.lower() in INPUT_SKIP_NAMES:
        return {}
    out: dict[str, InputRow] = {}
    for row in read_input(path):
        kn = normalize_kn(row.cadastral)
        if kn:
            out[kn] = row
    return out


def load_input_index(paths: list[Path]) -> dict[str, InputRow]:
    """Объединяет несколько входных файлов; непустые поля не затираются."""
    idx: dict[str, InputRow] = {}
    for p in paths:
        for kn, row in input_rows_by_kn(p).items():
            if kn not in idx:
                idx[kn] = row
            else:
                apply_input_to_row(idx[kn], row)
    return idx


def apply_input_to_row(row: InputRow, *sources: InputRow | dict | None):
    """Дополняет InputRow полями из кэша/входного файла/старых строк отчёта."""
    for src in sources:
        if not src:
            continue
        if isinstance(src, dict):
            if not row.ext_number:
                row.ext_number = _str(src.get("ext_number"))
            if not row.full_address:
                row.full_address = _str(src.get("full_address") or src.get("address"))
        else:
            if not row.ext_number and src.ext_number:
                row.ext_number = src.ext_number
            if not row.full_address and src.full_address:
                row.full_address = src.full_address


def load_api_cache(cache_dir: Path, kn: str) -> dict | None:
    """Загружает info из кэша, если файл есть."""
    path = cache_json_path(cache_dir, kn)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    info = data.get("info")
    return info if isinstance(info, dict) else None


# --------------------------------------------------------------------------- #
# Форма собственности (lk.rosreestr.ru) — отдельный бесплатный источник
# --------------------------------------------------------------------------- #
# API Контур.Реестро НЕ возвращает «Форму собственности» (напр. «Государственная
# федеральная»). По правке заказчика её берут со страницы
# https://lk.rosreestr.ru/eservices/real-estate-objects-online (строка
# «Форма собственности»). Этот сервис защищён капчей и агрессивно блокирует
# автоматизацию, поэтому источник реализован в три слоя с приоритетом:
#   1) ручной override-файл (input/ownership_forms.csv|json) — KN → форма;
#   2) локальный кэш output/cache/rosreestr/{kn}.json (одна запись на КН);
#   3) best-effort авто-запрос к lk.rosreestr.ru (включается флагом, требует
#      проверки/решения капчи на машине исполнителя).
# Если форму получить не удалось — в отчёт пишется «данные отсутствуют».
ROSREESTR_ONLINE_URL = "https://lk.rosreestr.ru/eservices/real-estate-objects-online"
OWNERSHIP_FORM_RE = re.compile(
    r"Форма\s+собственности\s*[:\-]?\s*(.+)", re.IGNORECASE
)


def ownership_cache_path(cache_dir: Path, kn: str) -> Path:
    return cache_dir / f"{sanitize_filename(normalize_kn(kn))}.json"


def load_ownership_overrides(path) -> dict:
    """Ручной справочник KN → форма собственности (CSV с ';' или JSON-объект)."""
    overrides: dict[str, str] = {}
    if not path:
        return overrides
    p = Path(path)
    if not p.exists():
        return overrides
    if p.suffix.lower() == ".json":
        with open(p, encoding="utf-8") as f:
            data = json.load(f) or {}
        for k, v in data.items():
            kn = normalize_kn(k)
            if kn:
                overrides[kn] = _str(v)
    else:
        with open(p, encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f, delimiter=";"):
                if len(row) >= 2 and CADASTRAL_RE.search(row[0] or ""):
                    overrides[normalize_kn(row[0])] = _str(row[1])
    return overrides


def load_ownership_cache(cache_dir: Path, kn: str) -> str | None:
    path = ownership_cache_path(cache_dir, kn)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return _str(data.get("ownership_form"))


def save_ownership_cache(cache_dir: Path, kn: str, form: str, source: str) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    path = ownership_cache_path(cache_dir, kn)
    payload = {
        "cadastralNumber": normalize_kn(kn),
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "ownership_form": _str(form),
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


def fetch_ownership_form_rosreestr(kn: str, timeout: int = 30) -> str:
    """
    Best-effort запрос «Формы собственности» к lk.rosreestr.ru по КН.

    ВНИМАНИЕ: публичный сервис Росреестра защищён капчей и блокирует
    автоматические обращения. Функция написана так, чтобы при любой ошибке
    (недоступность, капча, изменение разметки) вернуть пустую строку, не
    прерывая обработку. Полноценная авто-загрузка требует решения капчи
    (ручная сессия / антикапча-сервис) и проверки на машине исполнителя.
    """
    try:
        resp = requests.get(
            ROSREESTR_ONLINE_URL,
            params={"cadastralNumber": normalize_kn(kn)},
            timeout=timeout,
            headers={"User-Agent": "Mozilla/5.0"},
        )
    except Exception:
        return ""
    if resp.status_code != 200:
        return ""
    m = OWNERSHIP_FORM_RE.search(resp.text or "")
    if not m:
        return ""
    return re.split(r"[<\n\r]", m.group(1).strip())[0].strip()


def get_ownership_form(kn: str, cache_dir: Path, overrides: dict,
                       *, fetch: bool = False) -> str:
    """Возвращает форму собственности по приоритету override → кэш → авто-запрос."""
    kn_norm = normalize_kn(kn)
    if not kn_norm:
        return ""

    if kn_norm in overrides:
        form = overrides[kn_norm]
        save_ownership_cache(cache_dir, kn_norm, form, "override")
        return form

    cached = load_ownership_cache(cache_dir, kn_norm)
    if cached is not None:
        return cached

    if fetch:
        form = fetch_ownership_form_rosreestr(kn_norm)
        save_ownership_cache(cache_dir, kn_norm, form, "rosreestr")
        return form

    return ""


def patch_report_rights_columns(rows: list[list]) -> list[list]:
    """
    Заполняет пустые столбцы группы «Собственник» значением «данные отсутствуют».
    Индексы 28 (ФИО), 29 (Вид собственности), 31 (Право).
    Индекс 30 (Доля) НЕ трогаем: по правке заказчика пустая доля остаётся пустой.
    """
    out = []
    for row in rows:
        r = list(row[:35])
        while len(r) < 35:
            r.append("")
        for i in (28, 29, 31):
            if not _str(r[i]):
                r[i] = "данные отсутствуют"
        out.append(r)
    return out


def load_existing_report(out_path: Path, pdf_dir: Path) -> tuple[set[str], list[list]]:
    """
    Читает уже сохранённый report.xlsx.
    Объект считается обработанным, если в отчёте есть КН, имя PDF и файл PDF существует.
    """
    if not out_path.exists():
        return set(), []

    wb = load_workbook(out_path, read_only=True, data_only=True)
    # Лист данных, а не «Сводка»: wb.active может указывать на вкладку,
    # выбранную при последнем сохранении в Excel.
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    processed: set[str] = set()
    rows: list[list] = []

    for raw in ws.iter_rows(min_row=3, values_only=True):
        row = list(raw[:35])
        while len(row) < 35:
            row.append("")
        rows.append(row)
        kn = normalize_kn(row[1])
        pdf_name = _str(row[34])
        if kn and pdf_name.lower().endswith(".pdf") and (pdf_dir / pdf_name).exists():
            processed.add(kn)

    wb.close()
    return processed, rows


def _drop_kn_rows(rows: list[list], kn: str) -> list[list]:
    """Удаляет из отчёта все строки с указанным КН (для --force)."""
    target = normalize_kn(kn)
    return [r for r in rows if normalize_kn(r[1]) != target]


def write_report_xlsx(all_xlsx_rows: list[list], stats: dict, out_path: Path):
    """Сводный Excel в формате образца TZ/Пример_Ответ/report.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    H1 = [
        "EXT_NUMBER", "Кадастровый номер", "Адрес в свободной форме",
        "Регион", None, "Район", None, "Город", None,
        "Населенный пункт", None, "Улица", None, "Дом", None,
        "Корпус", None, "Строение", None, "Помещение", None,
        "Площадь, м2", "Кадастровая стоимость", "Стоит на учёте",
        "Дата снятия с учёта", "Категория земель",
        "Разрешённое использование", "Разрешённое использование по документам",
        "Собственник", None, None, None,
        "Номер выписки", "Дата выписки", "Отчёт",
    ]
    H2 = [
        None, None, None,
        "Тип", "Название", "Тип", "Название", "Тип", "Название",
        "Тип", "Название", "Тип", "Название", "Тип", "Номер",
        "Тип", "Номер", "Тип", "Номер", "Тип", "Номер",
        None, None, None, None, None, None, None,
        "ФИО", "Вид собственности", "Доля", "Право",
        None, None, None,
    ]
    assert len(H1) == 35 and len(H2) == 35

    body_font = Font(name="Calibri", size=12)
    ws.append(H1)
    ws.append(H2)
    for row_i in range(1, 3):
        for col_i in range(1, 36):
            cell = ws.cell(row=row_i, column=col_i)
            cell.font = body_font

    for row_data in all_xlsx_rows:
        ws.append(row_data[:35])

    for col_i in range(1, 36):
        ws.column_dimensions[get_column_letter(col_i)].width = 13

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=35):
        for cell in row:
            cell.font = body_font

    # ---- Лист «Сводка» (служебный; в образце его нет) ----
    ws2 = wb.create_sheet("Сводка")
    ws2["A1"] = "Отчёт о выполненных работах (ЕГРН)"
    ws2["A1"].font = Font(name="Calibri", size=12)
    ws2["A3"] = "Дата формирования"
    ws2["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws2["A4"] = "Всего объектов в задании"
    ws2["B4"] = stats["total"]
    ws2["A5"] = "Сформировано отчётов (данные ЕГРН получены)"
    ws2["B5"] = stats["ok"]
    ws2["A6"] = "Объектов без сведений (ошибка API / не найдено в ЕГРН)"
    ws2["B6"] = stats["failed"]
    ws2["A7"] = "Пропущено (нет кадастрового номера, PDF не создавался)"
    ws2["B7"] = stats.get("skipped_no_kn", 0)
    ws2["A8"] = "Пропущено (уже обработано ранее, без запроса к API)"
    ws2["B8"] = stats.get("skipped_already", 0)
    if stats.get("network_skipped"):
        ws2["A9"] = "Пропущено (обрыв сети, повторите запуск)"
        ws2["B9"] = stats.get("network_skipped", 0)
        summary_rows = 10
    else:
        summary_rows = 9
    for r in range(1, summary_rows):
        for c in range(1, 3):
            ws2.cell(row=r, column=c).font = body_font
    ws2.column_dimensions["A"].width = 52
    ws2.column_dimensions["B"].width = 20

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Основной сценарий
# --------------------------------------------------------------------------- #
def process(args):
    cfg    = load_config(Path(args.config))
    client = ReestroClient(
        cfg,
        pause=getattr(args, "pause", 0.4),
        timeout=getattr(args, "timeout", 120),
        retries=getattr(args, "retries", 5),
    )

    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Входной файл не найден: {input_path}")

    out_dir  = Path(args.output)
    pdf_dir  = out_dir / "pdf"
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)

    use_cache = not getattr(args, "no_cache", False)
    cache_dir = Path(args.cache_dir) if getattr(args, "cache_dir", None) else out_dir / "cache" / "json"
    if use_cache:
        cache_dir.mkdir(parents=True, exist_ok=True)

    # Форма собственности (lk.rosreestr.ru): override-файл + кэш + опц. авто-запрос
    rr_cache_dir = out_dir / "cache" / "rosreestr"
    rr_cache_dir.mkdir(parents=True, exist_ok=True)
    overrides_path = getattr(args, "ownership_forms", None) or (BASE_DIR / "input" / "ownership_forms.csv")
    ownership_overrides = load_ownership_overrides(overrides_path)
    fetch_rr = getattr(args, "fetch_rosreestr", False)
    if ownership_overrides:
        print(f"Форма собственности: загружено {len(ownership_overrides)} записей из {Path(overrides_path).name}")
    if fetch_rr:
        print("Форма собственности: включён авто-запрос к lk.rosreestr.ru (best-effort)")

    def ownership_for(kn: str) -> str:
        return get_ownership_form(kn, rr_cache_dir, ownership_overrides, fetch=fetch_rr)

    summary_path = out_dir / "report.xlsx"
    force = getattr(args, "force", False)
    processed_kn, existing_rows = load_existing_report(summary_path, pdf_dir)
    if processed_kn:
        print(f"Уже обработано (в {summary_path.name}): {len(processed_kn)} КН")

    items = read_input(input_path)
    print(f"Объектов во входном файле: {len(items)}")

    new_xlsx_rows: list[list] = []
    ok = failed = skipped_no_kn = skipped_already = network_skipped = 0
    today = datetime.now().strftime("%d.%m.%Y")
    limit = getattr(args, "limit", None)
    save_every = getattr(args, "save_every", 5)
    if save_every is not None and save_every < 1:
        save_every = 0
    if limit:
        print(f"Лимит новых объектов (API): {limit}")
    if save_every:
        print(f"Сохранение report.xlsx каждые {save_every} новых объект(ов)")
    print(f"API: timeout={client.timeout}с, повторов при обрыве={client.retries}, "
          f"пауза={client.pause}с")

    def save_progress(force: bool = False):
        nonlocal existing_rows, new_xlsx_rows
        if not new_xlsx_rows and not force:
            return
        merged = patch_report_rights_columns(existing_rows + new_xlsx_rows)
        write_report_xlsx(merged, {
            "total": len(items),
            "ok": ok,
            "failed": failed,
            "skipped_no_kn": skipped_no_kn,
            "skipped_already": skipped_already,
            "network_skipped": network_skipped,
        }, summary_path)
        existing_rows = merged
        new_xlsx_rows = []

    objects_since_save = 0

    def bump_save():
        nonlocal objects_since_save
        if not save_every:
            return
        objects_since_save += 1
        if objects_since_save >= save_every:
            save_progress()
            objects_since_save = 0

    for n, row in enumerate(items, start=1):
        cadastral = ""
        comment   = ""
        strategy  = "ЕГРН"

        # ---- 1. Кадастровый номер из входных данных ----
        m = CADASTRAL_RE.search(row.cadastral) if row.cadastral else None
        if m:
            cadastral = m.group(0)
        else:
            comment = "Кадастровый номер отсутствует во входном файле"

        label = row.cadastral or row.full_address or f"object_{n}"
        print(f"[{n}/{len(items)}] [{strategy}] {label[:60]} ...", end=" ")

        if not cadastral:
            skipped_no_kn += 1
            reason = comment or "кадастровый номер не определён"
            row.cadastral = cadastral
            xlsx_rows = build_xlsx_rows(row, None, [], "", "", today)
            new_xlsx_rows.extend(xlsx_rows)
            print(f"пропуск — {reason[:50]} (PDF не формируется)")
            continue

        row.cadastral = cadastral
        kn_norm = normalize_kn(cadastral)

        if kn_norm in processed_kn and not force:
            skipped_already += 1
            pdf_ref = next(
                (_str(r[34]) for r in existing_rows if normalize_kn(r[1]) == kn_norm and r[34]),
                "",
            )
            print(f"пропуск — уже обработан ({pdf_ref or 'report.xlsx'})")
            continue

        if force and kn_norm in processed_kn:
            existing_rows = _drop_kn_rows(existing_rows, kn_norm)
            processed_kn.discard(kn_norm)

        extract_id = new_extract_id()
        extract_date = today
        pdf_name = pdf_name_for_extract(extract_id)
        pdf_path = pdf_dir / pdf_name

        info = None
        from_cache = False
        if use_cache and not force:
            info = load_api_cache(cache_dir, kn_norm)
            if info:
                from_cache = True

        if from_cache:
            save_api_cache(cache_dir, kn_norm, info, None, input_row=row)
            rights = extract_rights(info)
            own_form = ownership_for(kn_norm)
            generate_pdf(info, rights, row, pdf_path, ownership_form=own_form)
            ok += 1
            xlsx_rows = build_xlsx_rows(
                row, info, rights, pdf_name, extract_id, extract_date,
                ownership_form=own_form)
            new_xlsx_rows.extend(xlsx_rows)
            processed_kn.add(kn_norm)
            print(f"OK из кэша (прав: {len(rights)})")
            bump_save()
            if limit and (ok + failed) >= limit:
                print(f"\nЛимит --limit {limit} достигнут, остановка.")
                break
            continue

        # ---- Запрос сведений по КН ----
        try:
            resp = client.object_info(cadastral)
        except RequestException as exc:
            network_skipped += 1
            print(f"сеть: {exc.__class__.__name__} — пропуск "
                  f"(объект не записан, повторите запуск)")
            continue

        if resp.status_code == 200:
            try:
                info = resp.json()
            except ValueError:
                info = None

            if info and info.get("cadastralNumber"):
                if use_cache:
                    path = save_api_cache(cache_dir, kn_norm, info, resp, input_row=row)
                    print(f"[cache: {path.name}] ", end="")
                rights = extract_rights(info)
                own_form = ownership_for(kn_norm)
                generate_pdf(info, rights, row, pdf_path, ownership_form=own_form)
                ok += 1
                xlsx_rows  = build_xlsx_rows(
                    row, info, rights, pdf_name, extract_id, extract_date,
                    ownership_form=own_form)
                new_xlsx_rows.extend(xlsx_rows)
                processed_kn.add(kn_norm)
                r_count = len(rights)
                print(f"OK  (прав: {r_count})"
                      + (f"  [{comment}]" if comment else ""))
                bump_save()
                if limit and (ok + failed) >= limit:
                    print(f"\nЛимит --limit {limit} достигнут, остановка.")
                    break
                continue
            reason = "пустой ответ (объект без КН)"
        elif resp.status_code == 404:
            reason = "объект не найден в ЕГРН (HTTP 404)"
        elif resp.status_code == 401:
            reason = "ошибка авторизации (HTTP 401) — проверьте apiKey/orgId"
        elif resp.status_code == 402:
            reason = ("недостаточно средств на балансе (HTTP 402) — "
                      "пополните balance для address_api_open_data")
        elif resp.status_code == 400:
            try:
                err = resp.json()
            except ValueError:
                err = {}
            msg = err.get("message", "")
            code = err.get("code", "")
            if code == "validation" or "тело" in msg.lower() or "баланс" in msg.lower():
                reason = (
                    "нет доступных единиц для address_api_open_data (HTTP 400). "
                    "Пополните предоплатный баланс в личном кабинете Контур.Реестро."
                )
            else:
                reason = f"некорректный запрос (HTTP 400) — {msg}"
        else:
            reason = f"ошибка API: HTTP {resp.status_code}"
            try:
                err = resp.json()
                if isinstance(err, dict) and err.get("message"):
                    reason += f" — {err['message']}"
            except ValueError:
                pass

        failed += 1
        full_reason = "; ".join(x for x in (comment, reason) if x)
        if not extract_id:
            extract_id = new_extract_id()
            extract_date = today
            pdf_name = pdf_name_for_extract(extract_id)
            pdf_path = pdf_dir / pdf_name
        generate_pdf_not_found(row, cadastral, full_reason, pdf_path)
        xlsx_rows = build_xlsx_rows(row, None, [], pdf_name, extract_id, extract_date)
        new_xlsx_rows.extend(xlsx_rows)
        processed_kn.add(kn_norm)
        print(full_reason[:80])
        bump_save()
        if limit and (ok + failed) >= limit:
            print(f"\nЛимит --limit {limit} достигнут, остановка.")
            break

    save_progress(force=True)

    stats = {
        "total": len(items),
        "ok": ok,
        "failed": failed,
        "skipped_no_kn": skipped_no_kn,
        "skipped_already": skipped_already,
        "network_skipped": network_skipped,
    }
    write_report_xlsx(
        patch_report_rights_columns(existing_rows + new_xlsx_rows),
        stats, summary_path)

    print("-" * 64)
    print(f"Готово.")
    print(f"  Объектов: {len(items)}")
    print(f"  Отчётов сформировано: {ok}")
    print(f"  Без сведений (ошибка API / не найдено): {failed}")
    if network_skipped:
        print(f"  Пропущено (обрыв сети, не записано — перезапустите): {network_skipped}")
    print(f"  Пропущено (нет КН, PDF не создавался): {skipped_no_kn}")
    print(f"  Пропущено (уже обработано): {skipped_already}")
    print(f"  PDF: {pdf_dir}")
    print(f"  Excel: {summary_path}")
    if use_cache:
        print(f"  Кэш JSON: {cache_dir}")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def main():
    p = argparse.ArgumentParser(
        description="Парсер открытых сведений ЕГРН через Контур.Реестро (address-api v2)."
    )
    p.add_argument(
        "-i", "--input",
        default=str(BASE_DIR / "TZ" / "Запрос.xlsx"),
        help="Входной XLSX/CSV. По умолчанию: TZ/Запрос.xlsx",
    )
    p.add_argument(
        "-o", "--output",
        default=str(BASE_DIR / "output"),
        help="Папка для результатов (PDF + report.xlsx).",
    )
    p.add_argument(
        "-c", "--config",
        default=str(BASE_DIR / "config.json"),
        help="Путь к config.json.",
    )
    p.add_argument(
        "--pause", type=float, default=2.0,
        help="Пауза между запросами к API, сек. (default: 2.0)",
    )
    p.add_argument(
        "--timeout", type=int, default=120, metavar="SEC",
        help="Таймаут ожидания ответа API, сек. (default: 120)",
    )
    p.add_argument(
        "--retries", type=int, default=5, metavar="N",
        help="Повторов запроса при обрыве/таймауте (default: 5)",
    )
    p.add_argument(
        "--save-every", type=int, default=5, metavar="N",
        help="Сохранять report.xlsx каждые N новых объектов (0 = только в конце, default: 5)",
    )
    p.add_argument(
        "--force", action="store_true",
        help="Повторно обработать КН, уже присутствующие в report.xlsx (новый PDF и API).",
    )
    p.add_argument(
        "--limit", type=int, default=None, metavar="N",
        help="Обработать не более N новых объектов через API (уже готовые пропускаются).",
    )
    p.add_argument(
        "--cache-dir",
        default=None,
        help="Папка для JSON-кэша ответов API (по умолчанию: {output}/cache/json).",
    )
    p.add_argument(
        "--no-cache",
        action="store_true",
        help="Не читать и не сохранять JSON-кэш ответов API.",
    )
    p.add_argument(
        "--ownership-forms",
        default=None,
        metavar="FILE",
        help="Ручной справочник «Форма собственности» (CSV ';' или JSON: КН → форма). "
             "По умолчанию: input/ownership_forms.csv (если есть).",
    )
    p.add_argument(
        "--fetch-rosreestr",
        action="store_true",
        help="Пытаться получить «Форму собственности» с lk.rosreestr.ru (best-effort, "
             "сервис с капчей — проверяйте результат).",
    )
    process(p.parse_args())


if __name__ == "__main__":
    main()
