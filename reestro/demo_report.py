# -*- coding: utf-8 -*-
"""Демо: PDF + Excel из mock-данных образца (5 прав)."""
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE))

from openpyxl import load_workbook

from fixtures_sample import SAMPLE_FIAS, SAMPLE_INFO, SAMPLE_RIGHTS
from reestro_parser import InputRow, build_xlsx_rows, generate_pdf, new_extract_id, pdf_name_for_extract, write_report_xlsx

OUT = Path(r"d:\Project_freelance\parser\demo_output")
PDF_DIR = OUT / "pdf"
PDF_DIR.mkdir(parents=True, exist_ok=True)

row = InputRow()
row.ext_number = "4700000043"
row.cadastral = SAMPLE_INFO["cadastralNumber"]
row.full_address = SAMPLE_INFO["address"]
row.fias_guid = SAMPLE_FIAS

extract_id = new_extract_id()
extract_date = "19.05.2026"
pdf_name = pdf_name_for_extract(extract_id)
pdf_path = PDF_DIR / pdf_name

# 1) PDF — один файл на объект
generate_pdf(SAMPLE_INFO, SAMPLE_RIGHTS, row, pdf_path)

# 2) Excel — одна строка на каждое право
rows = build_xlsx_rows(row, SAMPLE_INFO, SAMPLE_RIGHTS, pdf_name, extract_id, extract_date)
write_report_xlsx(
    rows,
    {"total": 1, "ok": 1, "failed": 0, "skipped_no_kn": 0},
    OUT / "report.xlsx",
)

print("PDF:", pdf_path)
print("Excel:", OUT / "report.xlsx")
print("Строк в Excel (прав):", len(rows))

# Сравнение с образцом
sample_path = next(BASE.glob("TZ/**/report.xlsx"))
sample_wb = load_workbook(sample_path, data_only=True)
demo_wb = load_workbook(OUT / "report.xlsx", data_only=True)
sw, dw = sample_wb.active, demo_wb.active

print("\n=== Сравнение с образцом ===")
print("Образец:", sample_path)
print("Колонок образец / наш:", sw.max_column, dw.max_column)
print("Ширина col A образец / наш:",
      sample_wb.active.column_dimensions["A"].width,
      demo_wb.active.column_dimensions["A"].width)

for c in range(1, 36):
    if sw.cell(1, c).value != dw.cell(1, c).value:
        print(f"  H1 col {c}: {sw.cell(1,c).value!r} vs {dw.cell(1,c).value!r}")
    if sw.cell(2, c).value != dw.cell(2, c).value:
        print(f"  H2 col {c}: {sw.cell(2,c).value!r} vs {dw.cell(2,c).value!r}")

# первая строка образца (право 1/5, reg 2023-24)
for r in range(3, sw.max_row + 1):
    if sw.cell(r, 32).value and "2023-24" in str(sw.cell(r, 32).value):
        sr = r
        break
else:
    sr = 3

dr = 3
cols_check = {
    1: "EXT_NUMBER",
    2: "КН",
    3: "Адрес",
    29: "Собственник",
    30: "Вид собственности",
    31: "Доля",
    32: "Право",
    33: "Номер выписки",
    34: "Дата выписки",
    35: "Отчёт",
}
print("\nСтрока образца (row", sr, ") vs наша (row", dr, "):")
for c, name in cols_check.items():
    sv, dv = sw.cell(sr, c).value, dw.cell(dr, c).value
    ok = "OK" if (sv == dv or (sv and dv and str(sv)[:20] == str(dv)[:20])) else "DIFF"
    if c in (32, 3):
        ok = "OK" if sv and dv and "43:31:070604:232" in str(sv) and "43:31:070604:232" in str(dv) else ok
    print(f"  col {c:2} {name:18} [{ok}]")
    if ok == "DIFF":
        print(f"       образец: {sv!r}")
        print(f"       наш:     {dv!r}")

print("\nСтрок с тем же PDF в образце:",
      sum(1 for r in range(3, sw.max_row + 1)
          if sw.cell(r, 35).value == "019e3f13-bbe0-73b3-818e-0aa9bbcce8d0.pdf"))
print("Строк с тем же PDF у нас:", len(rows))

sample_wb.close()
demo_wb.close()
