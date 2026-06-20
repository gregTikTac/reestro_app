# -*- coding: utf-8 -*-
"""Переименовать PDF и обновить report.xlsx без повторного запроса к API."""
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from reestro_parser import new_extract_id, pdf_name_for_extract

OUT = Path(__file__).resolve().parent / "output_test_first"
PDF_DIR = OUT / "pdf"
XLSX = OUT / "report.xlsx"
EXT_NUMBER = "77001001037"

old_name = "test-first-object.pdf"
old_pdf = PDF_DIR / old_name
if not old_pdf.exists():
    for p in PDF_DIR.glob("*.pdf"):
        if p.name != old_name:
            old_pdf = p
            old_name = p.name
            break

if not old_pdf.exists():
    raise SystemExit(f"PDF не найден в {PDF_DIR}")

wb = load_workbook(XLSX)
ws = wb.active
extract_date = ws.cell(3, 34).value
if not extract_date:
    extract_date = "17.06.2026"

extract_id = new_extract_id()
new_name = pdf_name_for_extract(extract_id)
new_pdf = PDF_DIR / new_name

if old_pdf.resolve() != new_pdf.resolve():
    if new_pdf.exists():
        new_pdf.unlink()
    old_pdf.rename(new_pdf)
    print(f"PDF: {old_name} -> {new_name}")

for r in range(3, ws.max_row + 1):
    ws.cell(r, 1).value = EXT_NUMBER
    ws.cell(r, 2).value = ws.cell(r, 2).value or "77:01:0001001:1037"
    for c in range(4, 29):
        ws.cell(r, c).value = None
    ws.cell(r, 33).value = extract_id
    ws.cell(r, 34).value = extract_date
    ws.cell(r, 35).value = new_name

wb.save(XLSX)
print(f"Excel обновлён: {XLSX}")
print(f"EXT_NUMBER: {EXT_NUMBER}")
print(f"Номер выписки: {extract_id}")
print(f"Отчёт: {new_name}")
