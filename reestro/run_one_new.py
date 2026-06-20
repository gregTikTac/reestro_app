# -*- coding: utf-8 -*-
"""
run_one_new.py — тестовый запуск для ОДНОГО объекта, которого ещё нет в report.xlsx.

Что делает:
  1. Читает Запрос.xlsx и report.xlsx
  2. Находит первый КН из Запрос.xlsx, которого НЕТ в report.xlsx
  3. Запускает по нему запрос к API Контур.Реестро
  4. Генерирует PDF и ДОЗАПИСЫВАЕТ строку в существующий report.xlsx
     (старые данные не стираются)

Запуск:
    python run_one_new.py

    # или явно задать конкретный КН:
    python run_one_new.py --kn 43:40:000184:744

    # или первые N новых объектов:
    python run_one_new.py --count 5

Требования:
    - config.json с apiKey и orgId рядом со скриптом (или reestro_parser.py)
    - reestro_parser.py должен быть в той же папке
    - pip install requests openpyxl fpdf2
"""

import argparse
import re
import sys
from pathlib import Path

# ── Пути ────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent
ZAPROS_FILE = BASE_DIR / "TZ" / "Запрос.xlsx"
OUTPUT_DIR  = BASE_DIR / "output"
REPORT_FILE = OUTPUT_DIR / "report.xlsx"
CONFIG_FILE = BASE_DIR / "config.json"

CADASTRAL_RE = re.compile(r"\d{2}:\d{2}:\d{1,7}:\d+")

# ── Импорт из основного парсера ──────────────────────────────────────────────
sys.path.insert(0, str(BASE_DIR))
try:
    from reestro_parser import (
        load_config,
        ReestroClient,
        read_input,
        load_existing_report,
        normalize_kn,
        extract_rights,
        generate_pdf,
        build_xlsx_rows,
        patch_report_rights_columns,
        write_report_xlsx,
        save_api_cache,
        load_api_cache,
        get_ownership_form,
        load_ownership_overrides,
        new_extract_id,
        pdf_name_for_extract,
        generate_pdf_not_found,
        _str,
    )
except ImportError as e:
    sys.exit(
        f"Ошибка импорта reestro_parser: {e}\n"
        "Убедитесь, что reestro_parser.py находится рядом с этим скриптом."
    )

from openpyxl import load_workbook
from datetime import datetime


def get_kn_from_report(report_path: Path) -> set:
    """Возвращает множество КН, уже присутствующих в report.xlsx."""
    if not report_path.exists():
        return set()
    try:
        wb = load_workbook(report_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return set()
        hdrs = [str(h).strip().lower() if h else "" for h in rows[0]]
        kn_idx = next(
            (i for i, h in enumerate(hdrs) if "кадастров" in h or "егрн" in h),
            None
        )
        if kn_idx is None:
            return set()
        result = set()
        for r in rows[1:]:
            val = str(r[kn_idx]).strip() if (kn_idx < len(r) and r[kn_idx]) else ""
            if CADASTRAL_RE.search(val):
                result.add(normalize_kn(val))
        return result
    except Exception as ex:
        print(f"[warn] Не удалось прочитать {report_path}: {ex}")
        return set()


def find_new_rows(zapros_path: Path, kn_in_report: set) -> list:
    """
    Возвращает список InputRow из Запрос.xlsx, КН которых нет в report.
    Только строки с валидным кадастровым номером.
    """
    all_rows = read_input(zapros_path)
    new = []
    for row in all_rows:
        kn = normalize_kn(row.cadastral)
        if kn and kn not in kn_in_report:
            new.append(row)
    return new


def process_rows(input_rows: list, cfg: dict, output_dir: Path, report_path: Path):
    """Обрабатывает список InputRow, дозаписывает результат в report.xlsx."""
    client    = ReestroClient(cfg, pause=0.4)
    cache_dir = output_dir / "cache" / "json"
    rr_cache  = output_dir / "cache" / "rosreestr"
    pdf_dir   = output_dir / "pdf"
    cache_dir.mkdir(parents=True, exist_ok=True)
    rr_cache.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)

    # Загружаем существующий report (не стираем!)
    _, existing_rows = load_existing_report(report_path, pdf_dir)
    kn_done = {normalize_kn(_str(r[1])) for r in existing_rows if r[1]}

    overrides_path = BASE_DIR / "input" / "ownership_forms.csv"
    overrides = load_ownership_overrides(overrides_path) if overrides_path.exists() else {}

    today    = datetime.now().strftime("%d.%m.%Y")
    new_rows = []
    ok = failed = 0

    for row in input_rows:
        kn_norm = normalize_kn(row.cadastral)
        if not kn_norm:
            continue
        if kn_norm in kn_done:
            print(f"[skip] {row.cadastral} — уже в report.xlsx")
            continue

        print(f"\n→ Запрос: {row.cadastral}")

        # Проверяем кэш (не тратим единицы повторно)
        info = load_api_cache(cache_dir, kn_norm)
        from_cache = bool(info)
        if from_cache:
            print("  [cache] данные взяты из локального кэша")
        else:
            resp = client.object_info(row.cadastral)
            print(f"  HTTP {resp.status_code}")

            if resp.status_code == 200:
                try:
                    info = resp.json()
                except ValueError:
                    info = None

                if info and info.get("cadastralNumber"):
                    save_api_cache(cache_dir, kn_norm, info, resp, input_row=row)
                else:
                    info = None

        extract_id   = new_extract_id()
        extract_date = today
        pdf_name     = pdf_name_for_extract(extract_id)
        pdf_path     = pdf_dir / pdf_name

        if info:
            rights    = extract_rights(info)
            own_form  = get_ownership_form(kn_norm, rr_cache, overrides, fetch=False)
            generate_pdf(info, rights, row, pdf_path, ownership_form=own_form)
            xlsx_rows = build_xlsx_rows(
                row, info, rights, pdf_name, extract_id, extract_date,
                ownership_form=own_form
            )
            area  = _str(info.get("area") or "")
            aunit = _str(info.get("areaUnit") or "")
            print(f"  ✓ OK | площадь: {area} {aunit} | прав: {len(rights)}")
            ok += 1
        else:
            # Определяем причину ошибки
            if not from_cache:
                reason = {
                    404: "объект не найден в ЕГРН",
                    401: "ошибка авторизации — проверьте apiKey/orgId",
                    402: "недостаточно средств на балансе",
                    400: "нет доступных единиц address_api_open_data",
                }.get(resp.status_code, f"HTTP {resp.status_code}")
                try:
                    err = resp.json()
                    if isinstance(err, dict) and err.get("message"):
                        reason += f" — {err['message']}"
                except Exception:
                    pass
                print(f"  ✗ {reason}")
            else:
                reason = "пустой кэш"
            generate_pdf_not_found(row, row.cadastral, reason, pdf_path)
            xlsx_rows = build_xlsx_rows(row, None, [], pdf_name, extract_id, extract_date)
            failed += 1

        new_rows.extend(xlsx_rows)
        kn_done.add(kn_norm)

    # Дозаписываем — объединяем старые строки + новые
    merged = patch_report_rights_columns(existing_rows + new_rows)

    # Пересчитываем статистику (все строки, включая старые)
    all_kns = {normalize_kn(_str(r[1])) for r in merged if r[1]}
    stats = {
        "total":           len(all_kns),
        "ok":              ok,
        "failed":          failed,
        "skipped_no_kn":   0,
        "skipped_already": 0,
    }
    write_report_xlsx(merged, stats, report_path)

    print("\n" + "─" * 60)
    print(f"Новых обработано: {ok + failed}  (успешно: {ok}, ошибок: {failed})")
    print(f"Итого строк в report.xlsx: {len(merged)}")
    print(f"PDF → {pdf_dir}")
    print(f"report.xlsx → {report_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Тестовый запуск одного нового объекта из Запрос.xlsx → report.xlsx"
    )
    parser.add_argument(
        "--kn", default=None,
        help="Конкретный кадастровый номер (если не задан — берётся первый новый из Запрос.xlsx)"
    )
    parser.add_argument(
        "--count", type=int, default=1,
        help="Сколько новых объектов обработать (default: 1)"
    )
    parser.add_argument(
        "--zapros", default=str(ZAPROS_FILE),
        help=f"Путь к Запрос.xlsx (default: {ZAPROS_FILE})"
    )
    parser.add_argument(
        "--output", default=str(OUTPUT_DIR),
        help=f"Папка с output/report.xlsx (default: {OUTPUT_DIR})"
    )
    parser.add_argument(
        "--config", default=str(CONFIG_FILE),
        help=f"Путь к config.json (default: {CONFIG_FILE})"
    )
    args = parser.parse_args()

    output_dir  = Path(args.output)
    report_path = output_dir / "report.xlsx"
    zapros_path = Path(args.zapros)

    if not zapros_path.exists():
        sys.exit(f"Файл не найден: {zapros_path}")

    cfg = load_config(Path(args.config))

    # Какие КН уже в report
    kn_in_report = get_kn_from_report(report_path)
    print(f"КН уже в report.xlsx: {len(kn_in_report)}")

    if args.kn:
        # Режим: конкретный КН
        kn = args.kn.strip()
        if normalize_kn(kn) in kn_in_report:
            print(f"КН {kn} уже есть в report.xlsx. Используйте --force в основном парсере для перезаписи.")
            sys.exit(0)
        # Найдём соответствующую строку из Запрос.xlsx
        all_rows = read_input(zapros_path)
        match = next((r for r in all_rows if normalize_kn(r.cadastral) == normalize_kn(kn)), None)
        if not match:
            print(f"КН {kn} не найден в {zapros_path}. Проверьте формат.")
            sys.exit(1)
        target_rows = [match]
    else:
        # Режим: первые N новых
        new_rows = find_new_rows(zapros_path, kn_in_report)
        if not new_rows:
            print("Все КН из Запрос.xlsx уже присутствуют в report.xlsx.")
            sys.exit(0)
        target_rows = new_rows[:args.count]
        print(f"Новых КН в Запрос.xlsx: {len(new_rows)}, обрабатываем: {len(target_rows)}")
        for r in target_rows:
            print(f"  → {r.cadastral}  ({r.full_address})")

    print()
    process_rows(target_rows, cfg, output_dir, report_path)


if __name__ == "__main__":
    main()
