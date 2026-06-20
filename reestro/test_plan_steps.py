# -*- coding: utf-8 -*-
"""Локальные шаги плана тестирования (без API)."""
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE))

from openpyxl import load_workbook

from reestro_parser import (
    InputRow,
    build_xlsx_rows,
    extract_rights,
    parse_right_string,
    read_input,
    CADASTRAL_RE,
    generate_pdf,
)
from fixtures_sample import SAMPLE_INFO, SAMPLE_RIGHTS, SAMPLE_FIAS


def step1():
    print("=== ШАГ 1: чтение входного файла ===")
    path = BASE / "input" / "objects_example.csv"
    rows = read_input(path)
    print(f"Файл: {path.name}, строк: {len(rows)}")
    valid = 0
    skip = 0
    for i, row in enumerate(rows, 1):
        m = CADASTRAL_RE.search(row.cadastral) if row.cadastral else None
        kn = m.group(0) if m else None
        if kn:
            valid += 1
            status = "БУДЕТ обработан"
        else:
            skip += 1
            status = "ПРОПУСК (PDF не создаётся)"
        print(f"  {i}. КН={kn or '—'} | {status}")
    print(f"Итого: {valid} с КН, {skip} без КН\n")
    return valid, skip


def step2():
    print("=== ШАГ 2: разбор строк прав ===")
    sample = (
        "№ 43:31:070604:232-43/052/2023-24 от 26.10.2023 "
        "(Долевая собственность, 1/5)"
    )
    r = parse_right_string(sample)
    assert r["right_type"] == "Долевая собственность", r
    assert r["share"] == "1/5", r
    assert r["number"] == "43:31:070604:232-43/052/2023-24", r
    assert r["reg_date"] == "26.10.2023", r
    print("  parse_right_string: OK")

    mock_info = {
        "ownershipType": "Частная",
        "rightAndRestrictions": [
            {
                "right": sample,
                "restrictions": ["не зарегистрировано"],
                "rightOwnerType": "Частная собственность",
            },
        ],
    }
    rights = extract_rights(mock_info)
    assert len(rights) == 1
    assert rights[0]["share"] == "1/5"
    assert rights[0]["encumbrances"] == ["не зарегистрировано"]
    print("  extract_rights: OK\n")


def step3():
    print("=== ШАГ 3: генерация PDF (mock, без API) ===")
    row = InputRow()
    row.cadastral = SAMPLE_INFO["cadastralNumber"]
    row.fias_guid = SAMPLE_FIAS
    out = Path(r"d:\Project_freelance\parser\test_plan_mock.pdf")
    generate_pdf(SAMPLE_INFO, SAMPLE_RIGHTS, row, out)
    assert out.exists() and out.stat().st_size > 5000
    print(f"  PDF создан: {out.name} ({out.stat().st_size // 1024} KB)")

    import fitz
    doc = fitz.open(str(out))
    text = "".join(doc[i].get_text() for i in range(len(doc)))
    pages = len(doc)
    doc.close()
    assert "43:31:070604:232" in text
    assert text.count("1/5") == len(SAMPLE_RIGHTS), (
        f"ожидалось {len(SAMPLE_RIGHTS)} долей 1/5, найдено {text.count('1/5')}"
    )
    assert text.count("2.1") == len(SAMPLE_RIGHTS)
    assert "  о  " not in text
    print(f"  {len(SAMPLE_RIGHTS)} дольщиков (1/5), {pages} стр., двойных пробелов нет\n")


def step4():
    print("=== ШАГ 4: формирование Excel (mock) ===")
    row = InputRow()
    row.ext_number = "EXT-001"
    row.cadastral = "43:31:070604:232"
    row.full_address = "тестовый адрес"
    row.region_name = "Кировская область"

    info = {
        "cadastralNumber": "43:31:070604:232",
        "address": "адрес из API",
        "area": "48,90",
        "cadastralValue": "322 596,18",
        "estateObjectStatus": "Актуально",
    }
    rights = [
        {
            "owner_type": "Частная собственность",
            "right_type": "Долевая собственность",
            "share": "1/5",
            "number": "43:31:070604:232-43/052/2023-24",
            "reg_date": "26.10.2023",
            "full_right": "",
        },
        {
            "owner_type": "Частная собственность",
            "right_type": "Долевая собственность",
            "share": "2/5",
            "number": "43:31:070604:232-43/052/2023-23",
            "reg_date": "26.10.2023",
            "full_right": "",
        },
    ]
    rows_data = build_xlsx_rows(row, info, rights, "test.pdf", "req-123", "17.06.2026")
    assert len(rows_data) == 2, "должно быть 2 строки (2 права)"

    row_no_kn = InputRow()
    row_no_kn.full_address = "без КН"
    rows_skip = build_xlsx_rows(row_no_kn, None, [], "", "", "17.06.2026")
    assert rows_skip[0][-1] == "", "для пропуска PDF-имя пустое"

    from reestro_parser import write_report_xlsx
    out = Path(r"d:\Project_freelance\parser\test_plan_report.xlsx")
    write_report_xlsx(
        rows_data + rows_skip,
        {"total": 3, "ok": 1, "failed": 0, "skipped_no_kn": 1},
        out,
    )
    wb = load_workbook(out, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    data_rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()
    assert len(data_rows) == 3
    assert data_rows[0][1] == "43:31:070604:232"
    assert data_rows[0][30] == "1/5"
    assert data_rows[1][30] == "2/5"
    assert data_rows[2][34] in (None, ""), "для пропуска PDF-имя пустое"
    print(f"  Excel: {out.name}, строк данных: {len(data_rows)}")
    print("  2 права -> 2 строки с долями 1/5 и 2/5; строка без КН -> пустой PDF\n")


if __name__ == "__main__":
    step1()
    step2()
    step3()
    step4()
    print("=== ШАГ 5: реальный API — НЕ запускался ===")
    print("Следующий шаг: 1 объект с КН через Postman + python reestro_parser.py -i ...")
