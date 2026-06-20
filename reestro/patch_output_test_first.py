# -*- coding: utf-8 -*-
"""Обновить report.xlsx и создать JSON-кэш без запросов к API."""
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from reestro_parser import (
    cache_json_path,
    load_existing_report,
    normalize_kn,
    patch_report_rights_columns,
    save_api_cache,
    write_report_xlsx,
)

OUT = Path(__file__).resolve().parent / "output_test_first"
CACHE = OUT / "cache" / "json"
REPORT = OUT / "report.xlsx"
PDF_DIR = OUT / "pdf"

# Известные данные первого тестового объекта (уже получены ранее)
KNOWN = {
    "77:01:0001001:1037": {
        "cadastralNumber": "77:01:0001001:1037",
        "address": "город Москва, территория Кремль, здание 21",
        "rightAndRestrictions": [
            {
                "right": "№ 77-77-13/001/2005-940 от 04.05.2005 (Оперативное управление)",
                "restrictions": [],
            },
            {
                "right": "№ 77-77-13/001/2005-941 от 04.05.2005 (Собственность)",
                "restrictions": [],
            },
        ],
    },
}


class _FakeResp:
    status_code = 200
    headers = {}


def seed_cache(rows: list[list]):
    CACHE.mkdir(parents=True, exist_ok=True)
    seen = set()
    for row in rows:
        kn = normalize_kn(row[1])
        if not kn or kn in seen:
            continue
        seen.add(kn)
        path = cache_json_path(CACHE, kn)
        if path.exists():
            continue
        info = KNOWN.get(kn)
        if not info:
            info = {
                "cadastralNumber": kn,
                "address": _str(row[2]),
                "rightAndRestrictions": [],
            }
        save_api_cache(CACHE, kn, info, _FakeResp())
        print(f"кэш: {path.name}")


def _str(v):
    return "" if v is None else str(v).strip()


def main():
    processed, existing_rows = load_existing_report(REPORT, PDF_DIR)
    patched = patch_report_rights_columns(existing_rows)

    stats = {
        "total": len({normalize_kn(r[1]) for r in patched if normalize_kn(r[1])}),
        "ok": len(processed),
        "failed": 0,
        "skipped_no_kn": sum(1 for r in patched if not normalize_kn(r[1])),
        "skipped_already": 0,
    }
    write_report_xlsx(patched, stats, REPORT)
    print(f"report.xlsx обновлён: {len(patched)} строк данных")

    seed_cache(patched)

    wb = load_workbook(REPORT, data_only=True)
    ws = wb.active
    print("\nПроверка AC–AF (строка 5, KN 43:40:000028:599):")
    for c in range(29, 33):
        print(f"  col {c}: {ws.cell(5, c).value!r}")
    wb.close()


if __name__ == "__main__":
    main()
